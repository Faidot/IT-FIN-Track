"""
Reports and Audit Trail views for IT FIN Track.
Enhanced version with detailed breakdowns and statistics.
"""

import io
from datetime import datetime, timedelta
from decimal import Decimal
from calendar import monthrange

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Max, Min
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
from django.utils import timezone
from django.core.paginator import Paginator

from core.models import Income, Expense, AuditLog, Category, Vendor, IncomeSource, User


@login_required
def report_dashboard(request):
    """Enhanced reports dashboard with summary statistics."""
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    first_day_of_year = today.replace(month=1, day=1)
    
    # Overall totals (only approved expenses)
    total_income = Income.objects.filter(is_soft_deleted=False).aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    total_expense = Expense.objects.filter(is_soft_deleted=False, status='approved').aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    
    # Monthly totals (only approved expenses)
    monthly_income = Income.objects.filter(
        is_soft_deleted=False, date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    monthly_expense = Expense.objects.filter(
        is_soft_deleted=False, status='approved', date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Yearly totals (only approved expenses)
    yearly_income = Income.objects.filter(
        is_soft_deleted=False, date__gte=first_day_of_year
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    yearly_expense = Expense.objects.filter(
        is_soft_deleted=False, status='approved', date__gte=first_day_of_year
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Transaction counts (approved only for expenses)
    income_count = Income.objects.filter(is_soft_deleted=False).count()
    expense_count = Expense.objects.filter(is_soft_deleted=False, status='approved').count()
    
    # Pending approvals
    pending_approvals = Expense.objects.filter(
        is_soft_deleted=False, status='pending'
    ).count()
    
    # Pending reimbursements
    pending_reimbursements = Income.objects.filter(
        is_soft_deleted=False, is_reimbursable=True, reimbursed=False
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Top 5 categories this month (approved only)
    top_categories = Expense.objects.filter(
        is_soft_deleted=False, status='approved', date__gte=first_day_of_month
    ).values('category__name', 'category__color').annotate(
        total=Sum('amount')
    ).order_by('-total')[:5]
    
    # Monthly trend (last 6 months)
    six_months_ago = today - timedelta(days=180)
    monthly_trend_income = Income.objects.filter(
        is_soft_deleted=False, date__gte=six_months_ago
    ).annotate(month=TruncMonth('date')).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    monthly_trend_expense = Expense.objects.filter(
        is_soft_deleted=False, status='approved', date__gte=six_months_ago
    ).annotate(month=TruncMonth('date')).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    # Prepare chart data
    trend_labels = []
    trend_income = []
    trend_expense = []
    
    income_by_month = {item['month']: float(item['total']) for item in monthly_trend_income}
    expense_by_month = {item['month']: float(item['total']) for item in monthly_trend_expense}
    
    for i in range(5, -1, -1):
        month_date = (today - timedelta(days=i*30)).replace(day=1)
        trend_labels.append(month_date.strftime('%b %Y'))
        
        income_val = 0
        expense_val = 0
        for key, val in income_by_month.items():
            if key.year == month_date.year and key.month == month_date.month:
                income_val = val
                break
        for key, val in expense_by_month.items():
            if key.year == month_date.year and key.month == month_date.month:
                expense_val = val
                break
        
        trend_income.append(income_val)
        trend_expense.append(expense_val)
    
    # Category chart data
    cat_labels = [c['category__name'] or 'Other' for c in top_categories]
    cat_values = [float(c['total']) for c in top_categories]
    cat_colors = [c['category__color'] or '#FF6B01' for c in top_categories]
    
    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': total_income - total_expense,
        'monthly_income': monthly_income,
        'monthly_expense': monthly_expense,
        'monthly_balance': monthly_income - monthly_expense,
        'yearly_income': yearly_income,
        'yearly_expense': yearly_expense,
        'yearly_balance': yearly_income - yearly_expense,
        'income_count': income_count,
        'expense_count': expense_count,
        'pending_approvals': pending_approvals,
        'pending_reimbursements': pending_reimbursements,
        'top_categories': top_categories,
        'trend_labels': trend_labels,
        'trend_income': trend_income,
        'trend_expense': trend_expense,
        'cat_labels': cat_labels,
        'cat_values': cat_values,
        'cat_colors': cat_colors,
    }
    
    return render(request, 'core/reports/dashboard.html', context)


@login_required
def monthly_expense_report(request):
    """Enhanced monthly IT expense report."""
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    # Date range
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date()
    else:
        end_date = datetime(year, month + 1, 1).date()
    
    # Previous month for comparison
    if month == 1:
        prev_start = datetime(year - 1, 12, 1).date()
        prev_end = datetime(year, 1, 1).date()
    else:
        prev_start = datetime(year, month - 1, 1).date()
        prev_end = start_date
    
    # Current month expenses
    expenses = Expense.objects.filter(
        is_soft_deleted=False,
        date__gte=start_date,
        date__lt=end_date
    ).select_related('category', 'vendor', 'created_by')
    
    # Previous month expenses for comparison
    prev_expenses = Expense.objects.filter(
        is_soft_deleted=False,
        date__gte=prev_start,
        date__lt=prev_end
    )
    
    # Summary stats
    total_expense = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    prev_total = prev_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    change_percent = ((total_expense - prev_total) / prev_total * 100) if prev_total > 0 else Decimal('0')
    
    avg_expense = expenses.aggregate(avg=Avg('amount'))['avg'] or Decimal('0')
    max_expense = expenses.aggregate(max=Max('amount'))['max'] or Decimal('0')
    transaction_count = expenses.count()
    
    # Category breakdown with percentages
    category_breakdown = expenses.values(
        'category__name', 'category__color', 'category__icon'
    ).annotate(
        total=Sum('amount'),
        count=Count('id'),
        avg=Avg('amount')
    ).order_by('-total')
    
    # Add percentage to each category
    category_data = []
    for cat in category_breakdown:
        cat['percentage'] = (cat['total'] / total_expense * 100) if total_expense > 0 else 0
        category_data.append(cat)
    
    # Vendor breakdown
    vendor_breakdown = expenses.exclude(vendor__isnull=True).values(
        'vendor__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    # Daily breakdown
    daily_breakdown = expenses.annotate(
        day=TruncDate('date')
    ).values('day').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('day')
    
    # Approval status breakdown
    status_breakdown = expenses.values('status').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Weekly breakdown
    weekly_breakdown = expenses.annotate(
        week=TruncWeek('date')
    ).values('week').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('week')
    
    # Chart data
    cat_labels = [c['category__name'] or 'Other' for c in category_data]
    cat_values = [float(c['total']) for c in category_data]
    cat_colors = [c['category__color'] or '#FF6B01' for c in category_data]
    
    daily_labels = [d['day'].strftime('%d') for d in daily_breakdown]
    daily_values = [float(d['total']) for d in daily_breakdown]
    
    context = {
        'year': year,
        'month': month,
        'month_name': datetime(year, month, 1).strftime('%B'),
        'start_date': start_date,
        'end_date': end_date - timedelta(days=1),
        'expenses': expenses[:50],  # Limit for display
        'total_expense': total_expense,
        'prev_total': prev_total,
        'change_percent': change_percent,
        'avg_expense': avg_expense,
        'max_expense': max_expense,
        'transaction_count': transaction_count,
        'category_breakdown': category_data,
        'vendor_breakdown': vendor_breakdown,
        'daily_breakdown': daily_breakdown,
        'status_breakdown': status_breakdown,
        'weekly_breakdown': weekly_breakdown,
        'years': range(2020, timezone.now().year + 1),
        'months': range(1, 13),
        'cat_labels': cat_labels,
        'cat_values': cat_values,
        'cat_colors': cat_colors,
        'daily_labels': daily_labels,
        'daily_values': daily_values,
    }
    
    return render(request, 'core/reports/monthly_expense.html', context)


@login_required
def income_expense_statement(request):
    """Enhanced Income vs Expense statement."""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = timezone.now().strftime('%Y-%m-%d')
    
    # Get data
    incomes = Income.objects.filter(
        is_soft_deleted=False,
        date__gte=date_from,
        date__lte=date_to
    ).select_related('source', 'created_by').order_by('-date')
    
    expenses = Expense.objects.filter(
        is_soft_deleted=False,
        date__gte=date_from,
        date__lte=date_to
    ).select_related('category', 'vendor', 'created_by').order_by('-date')
    
    # Summary stats
    total_income = incomes.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_expense = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    net_balance = total_income - total_expense
    
    income_count = incomes.count()
    expense_count = expenses.count()
    
    avg_income = incomes.aggregate(avg=Avg('amount'))['avg'] or Decimal('0')
    avg_expense = expenses.aggregate(avg=Avg('amount'))['avg'] or Decimal('0')
    
    # Income by source with details
    income_by_source = incomes.values(
        'source__name', 'source__icon', 'source__color'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Add percentage
    income_source_data = []
    for src in income_by_source:
        src['percentage'] = (src['total'] / total_income * 100) if total_income > 0 else 0
        income_source_data.append(src)
    
    # Expense by category
    expense_by_category = expenses.values(
        'category__name', 'category__icon', 'category__color'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    expense_cat_data = []
    for cat in expense_by_category:
        cat['percentage'] = (cat['total'] / total_expense * 100) if total_expense > 0 else 0
        expense_cat_data.append(cat)
    
    # Daily trend
    daily_income = incomes.annotate(
        day=TruncDate('date')
    ).values('day').annotate(total=Sum('amount')).order_by('day')
    
    daily_expense = expenses.annotate(
        day=TruncDate('date')
    ).values('day').annotate(total=Sum('amount')).order_by('day')
    
    # Payment mode breakdown
    payment_modes = incomes.values('payment_mode').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Chart data
    inc_labels = [s['source__name'] or 'Other' for s in income_source_data]
    inc_values = [float(s['total']) for s in income_source_data]
    inc_colors = [s['source__color'] or '#28A745' for s in income_source_data]
    
    exp_labels = [c['category__name'] or 'Other' for c in expense_cat_data]
    exp_values = [float(c['total']) for c in expense_cat_data]
    exp_colors = [c['category__color'] or '#DC3545' for c in expense_cat_data]
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'incomes': incomes[:30],
        'expenses': expenses[:30],
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'income_count': income_count,
        'expense_count': expense_count,
        'avg_income': avg_income,
        'avg_expense': avg_expense,
        'income_by_source': income_source_data,
        'expense_by_category': expense_cat_data,
        'daily_income': daily_income,
        'daily_expense': daily_expense,
        'payment_modes': payment_modes,
        'inc_labels': inc_labels,
        'inc_values': inc_values,
        'inc_colors': inc_colors,
        'exp_labels': exp_labels,
        'exp_values': exp_values,
        'exp_colors': exp_colors,
    }
    
    return render(request, 'core/reports/income_expense.html', context)


@login_required
def reimbursement_report(request):
    """Recurring bill payment report by month."""
    from core.models import RecurringBill, BillPayment
    from datetime import date
    
    today = date.today()
    current_year = int(request.GET.get('year', today.year))
    
    # Get all active recurring bills
    bills = RecurringBill.objects.filter(
        is_soft_deleted=False,
        is_active=True
    ).select_related('category', 'vendor').prefetch_related('payments')
    
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    # Build month-wise payment matrix for each bill (track by period_start = billing month)
    bill_data = []
    for bill in bills:
        months = []
        for m in range(1, 13):
            # Check payment by billing period month (period_start)
            payment = bill.payments.filter(
                period_start__year=current_year,
                period_start__month=m,
                status='paid'
            ).first()
            
            pending_payment = bill.payments.filter(
                period_start__year=current_year,
                period_start__month=m,
                status='pending'
            ).first()
            
            months.append({
                'month': m,
                'month_name': month_names[m - 1],
                'paid': payment is not None,
                'pending': pending_payment is not None,
                'amount': payment.amount if payment else None,
                'payment_type': payment.payment_type if payment else None,
                'is_current': m == today.month and current_year == today.year,
                'is_future': (m > today.month and current_year == today.year) or current_year > today.year,
                'is_overdue': pending_payment.is_overdue if pending_payment else False
            })
        
        # Calculate totals for this bill
        total_paid = bill.payments.filter(
            period_end__year=current_year,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        paid_count = bill.payments.filter(
            period_end__year=current_year,
            status='paid'
        ).count()
        
        bill_data.append({
            'bill': bill,
            'months': months,
            'total_paid': total_paid,
            'paid_count': paid_count,
        })
    
    # Overall summary
    total_paid_year = BillPayment.objects.filter(
        bill__is_soft_deleted=False,
        period_end__year=current_year,
        status='paid'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    total_pending = BillPayment.objects.filter(
        bill__is_soft_deleted=False,
        status='pending'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # IT Payment vs Accounts Pay breakdown
    it_payment_total = BillPayment.objects.filter(
        bill__is_soft_deleted=False,
        period_end__year=current_year,
        status='paid',
        payment_type='it_payment'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    accounts_pay_total = BillPayment.objects.filter(
        bill__is_soft_deleted=False,
        period_end__year=current_year,
        status='paid',
        payment_type='accounts_pay'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Available years for navigation
    years = list(range(today.year - 2, today.year + 2))
    
    context = {
        'bill_data': bill_data,
        'month_names': month_names,
        'current_year': current_year,
        'current_month': today.month,
        'years': years,
        'total_paid_year': total_paid_year,
        'total_pending': total_pending,
        'it_payment_total': it_payment_total,
        'accounts_pay_total': accounts_pay_total,
    }
    
    return render(request, 'core/reports/reimbursement.html', context)


@login_required
def audit_trail(request):
    """Audit trail report - Admin only."""
    # Admin only
    if not request.user.is_admin:
        messages.error(request, 'Only administrators can access the audit trail.')
        return redirect('core:dashboard')
    
    logs = AuditLog.objects.select_related('user').all()
    
    # Filters
    user_id = request.GET.get('user', '')
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    action = request.GET.get('action', '')
    if action:
        logs = logs.filter(action=action)
    
    model = request.GET.get('model', '')
    if model:
        logs = logs.filter(model_name=model)
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)
    
    search = request.GET.get('search', '').strip()
    if search:
        logs = logs.filter(changes_summary__icontains=search)
    
    # Pagination
    paginator = Paginator(logs.order_by('-timestamp'), 50)
    page = request.GET.get('page', 1)
    logs = paginator.get_page(page)
    
    # Filter options
    users = User.objects.filter(is_soft_deleted=False)
    actions = AuditLog.ActionType.choices
    models = AuditLog.objects.values_list('model_name', flat=True).distinct()
    
    context = {
        'logs': logs,
        'users': users,
        'actions': actions,
        'models': models,
        'user_filter': user_id,
        'action_filter': action,
        'model_filter': model,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
    }
    
    return render(request, 'core/reports/audit_trail.html', context)


@login_required
def export_excel(request, report_type):
    """Export report to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse('openpyxl not installed', status=500)
    
    wb = Workbook()
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4E4E4E', end_color='4E4E4E', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if report_type == 'expenses':
        ws.title = 'Expenses'
        headers = ['Date', 'Category', 'Vendor', 'Amount', 'Description', 'Purpose', 'Status', 'Created By']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        expenses = Expense.objects.filter(is_soft_deleted=False).select_related(
            'category', 'vendor', 'created_by'
        ).order_by('-date')
        
        for row, expense in enumerate(expenses, 2):
            ws.cell(row=row, column=1, value=expense.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=expense.category.name if expense.category else '')
            ws.cell(row=row, column=3, value=expense.vendor.name if expense.vendor else '')
            ws.cell(row=row, column=4, value=float(expense.amount))
            ws.cell(row=row, column=5, value=expense.description)
            ws.cell(row=row, column=6, value=expense.purpose)
            ws.cell(row=row, column=7, value=expense.get_status_display())
            ws.cell(row=row, column=8, value=expense.created_by.get_full_name() or expense.created_by.username)
        
        filename = f'expenses_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    elif report_type == 'incomes':
        ws.title = 'Incomes'
        headers = ['Date', 'Source', 'Amount', 'Payment Mode', 'Reference', 'Description', 'Reimbursable', 'Created By']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        incomes = Income.objects.filter(is_soft_deleted=False).select_related(
            'source', 'created_by'
        ).order_by('-date')
        
        for row, income in enumerate(incomes, 2):
            ws.cell(row=row, column=1, value=income.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=income.source.name if income.source else '')
            ws.cell(row=row, column=3, value=float(income.amount))
            ws.cell(row=row, column=4, value=income.get_payment_mode_display())
            ws.cell(row=row, column=5, value=income.reference_number)
            ws.cell(row=row, column=6, value=income.description)
            ws.cell(row=row, column=7, value='Yes' if income.is_reimbursable else 'No')
            ws.cell(row=row, column=8, value=income.created_by.get_full_name() or income.created_by.username)
        
        filename = f'incomes_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    else:
        return HttpResponse('Invalid report type', status=400)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def account_balance_report(request):
    """Account Balance Report - Shows remaining balance by income source."""
    from django.db.models import F, Value, DecimalField
    from django.db.models.functions import Coalesce
    
    # Date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    source_filter = request.GET.get('source', '')
    
    # Base income queryset
    incomes = Income.objects.filter(is_soft_deleted=False).select_related('source')
    
    if date_from:
        incomes = incomes.filter(date__gte=date_from)
    if date_to:
        incomes = incomes.filter(date__lte=date_to)
    if source_filter:
        incomes = incomes.filter(source_id=source_filter)
    
    # Calculate spent for each income using annotation
    income_balances = []
    total_received = Decimal('0')
    total_spent = Decimal('0')
    total_remaining = Decimal('0')
    
    for income in incomes.order_by('-date'):
        spent = income.spent_amount
        remaining = income.remaining_amount
        total_received += income.amount
        total_spent += spent
        total_remaining += remaining
        
        income_balances.append({
            'income': income,
            'spent': spent,
            'remaining': remaining,
            'usage_percent': (spent / income.amount * 100) if income.amount > 0 else 0,
        })
    
    # Summary by Source
    source_summary = {}
    for item in income_balances:
        source_name = item['income'].source.name
        source_id = item['income'].source.id
        source_color = item['income'].source.color
        source_icon = item['income'].source.icon
        
        if source_name not in source_summary:
            source_summary[source_name] = {
                'id': source_id,
                'name': source_name,
                'color': source_color,
                'icon': source_icon,
                'received': Decimal('0'),
                'spent': Decimal('0'),
                'remaining': Decimal('0'),
                'count': 0,
            }
        
        source_summary[source_name]['received'] += item['income'].amount
        source_summary[source_name]['spent'] += item['spent']
        source_summary[source_name]['remaining'] += item['remaining']
        source_summary[source_name]['count'] += 1
    
    # Convert to list and add percentages
    source_data = []
    for data in source_summary.values():
        data['usage_percent'] = (data['spent'] / data['received'] * 100) if data['received'] > 0 else 0
        source_data.append(data)
    
    # Sort by remaining descending
    source_data.sort(key=lambda x: x['remaining'], reverse=True)
    
    # Get sources for filter dropdown
    sources = IncomeSource.objects.filter(is_soft_deleted=False, is_active=True)
    
    context = {
        'income_balances': income_balances,
        'source_data': source_data,
        'sources': sources,
        'total_received': total_received,
        'total_spent': total_spent,
        'total_remaining': total_remaining,
        'date_from': date_from,
        'date_to': date_to,
        'source_filter': source_filter,
    }
    
    return render(request, 'core/reports/account_balance.html', context)

